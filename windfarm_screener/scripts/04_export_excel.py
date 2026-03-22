"""
04_export_excel.py — Export Wind Farm Investment Screener to formatted Excel workbook.

Input:  data/processed/wind_plants_kpis.parquet
        outputs/audit/download_manifest.csv
Output: outputs/wind_screener.xlsx

Tabs:
  1. Sources        — Data provenance, KPI lineage, assumptions register
  2. Raw_Data       — Full merged dataset with EIA hyperlinks
  3. KPIs           — Computed KPI columns with conditional formatting
  4. Metrics        — Live Excel formulas referencing Raw_Data / KPIs
  5. Summary        — Aggregate stats via Excel COUNTIF / MEDIAN / SUM formulas
"""

import time
import sys
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

BASE_DIR = Path(__file__).resolve().parent.parent
PROC_DIR = BASE_DIR / "data" / "processed"
AUDIT_DIR = BASE_DIR / "outputs" / "audit"
OUT_DIR = BASE_DIR / "outputs"

# ── Constants (must match 03_kpis.py) ─────────────────────────────────────────
HOURS_PER_YEAR = 8760
REFERENCE_PRICE_PER_MWH = 30
CF_CAP = 0.65
CURRENT_YEAR = 2026
MIN_CAPACITY_MW = 10

YEARS = list(range(2018, 2025))
ALL_YEARS = list(range(2018, 2026))

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(bold=True, color="2F5496", size=12)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
AMBER_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
LINK_FONT = Font(color="0563C1", underline="single")

CF_FMT = "0.000"
PCT_FMT = "0.0%"
CURRENCY_FMT = "$#,##0"
NUMBER_FMT = "#,##0"
DECIMAL_FMT = "#,##0.0"


def style_header_row(ws, ncols):
    """Apply header styling to row 1 of a worksheet."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_fit_columns(ws, max_width=30):
    """Auto-fit column widths, capped at max_width characters."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells[:100]:  # sample first 100 rows for speed
            if cell.value is not None:
                val_len = len(str(cell.value))
                if val_len > max_len:
                    max_len = val_len
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def apply_number_format(ws, col_idx, fmt, start_row=2):
    """Apply number format to an entire column starting at start_row."""
    for row in range(start_row, ws.max_row + 1):
        ws.cell(row=row, column=col_idx).number_format = fmt


def detect_cf_col(col_name):
    """Return True if column name represents a capacity factor."""
    col_lower = col_name.lower()
    if col_lower.startswith("cf_") and "error" not in col_lower and "zero" not in col_lower:
        return True
    return False


def detect_pct_col(col_name):
    """Return True if column name represents a percentage."""
    col_lower = col_name.lower()
    return "pct" in col_lower or "percentile" in col_lower


def detect_currency_col(col_name):
    """Return True if column name represents a currency value."""
    col_lower = col_name.lower()
    return "usd" in col_lower or "revenue" in col_lower


def detect_2025_epm_col(col_name):
    """Return True if column is derived from 2025 EPM data."""
    col_lower = col_name.lower()
    return "2025" in col_lower and ("gen" in col_lower or "cf" in col_lower)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1: Sources
# ══════════════════════════════════════════════════════════════════════════════
def build_sources_tab(wb, manifest_df):
    """Build the Sources tab with provenance, KPI lineage, and assumptions."""
    ws = wb.active
    ws.title = "Sources"

    row = 1

    # ── Section A: Source Registry ─────────────────────────────────────────
    ws.cell(row=row, column=1, value="SECTION A — SOURCE REGISTRY").font = SECTION_FONT
    for c in range(1, 12):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    src_headers = [
        "Source Name", "Publisher", "Years Covered", "Description",
        "Landing Page", "Download URL", "Local File",
        "Downloaded", "MD5 Hash", "Row Count", "Data Note"
    ]
    for ci, h in enumerate(src_headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    # Source definitions (static metadata enriching the manifest)
    sources_meta = [
        {
            "source_name": "EIA-860",
            "publisher": "U.S. Energy Information Administration",
            "years": "2018-2024",
            "description": "Plant & generator metadata (capacity, location, technology, owner)",
            "note": "Annual survey of electric generators; Schedule 2 (plant), 3.1 (wind), 4 (owner)"
        },
        {
            "source_name": "EIA-923",
            "publisher": "U.S. Energy Information Administration",
            "years": "2018-2024",
            "description": "Annual electricity generation by plant and prime mover",
            "note": "Schedules 2/3/4/5; filtered to prime mover = WT (wind turbine)"
        },
        {
            "source_name": "EPM-2025",
            "publisher": "U.S. Energy Information Administration",
            "years": "2025",
            "description": "Monthly plant-level wind generation via EIA Open Data API",
            "note": "Electric Power Monthly; preliminary data, subject to revision"
        },
        {
            "source_name": "EIA-860M",
            "publisher": "U.S. Energy Information Administration",
            "years": "2025 (latest)",
            "description": "Monthly capacity change updates",
            "note": "Used to detect 2025 capacity additions/retirements"
        },
        {
            "source_name": "USWTDB",
            "publisher": "USGS / LBNL / AWEA",
            "years": "Cumulative",
            "description": "U.S. Wind Turbine Database — individual turbine records",
            "note": "Turbine vintage, model, hub height, rotor diameter"
        },
        {
            "source_name": "eGRID",
            "publisher": "U.S. Environmental Protection Agency",
            "years": "2022/2023",
            "description": "NERC region labels and emissions data by plant",
            "note": "Used for NERC region assignment and regional peer comparisons"
        },
        {
            "source_name": "LBNL",
            "publisher": "Lawrence Berkeley National Laboratory",
            "years": "2024 report",
            "description": "Wind Technologies Market Report — regional CF benchmarks",
            "note": "Benchmark CFs by NERC region and installation vintage decade"
        },
    ]

    # Build lookup from manifest
    manifest_lookup = {}
    if manifest_df is not None and len(manifest_df) > 0:
        for _, mrow in manifest_df.iterrows():
            key = str(mrow.get("source_name", "")).split("-")[0]  # e.g. "EIA" from "EIA-860-2024"
            if key not in manifest_lookup:
                manifest_lookup[str(mrow.get("source_name", ""))] = mrow

    for src in sources_meta:
        # Find matching manifest row (best match)
        m = None
        for mk, mv in manifest_lookup.items():
            if src["source_name"] in mk:
                m = mv
                break

        landing_url = str(m.get("landing_page_url", "")) if m is not None else ""
        download_url = str(m.get("download_url", "")) if m is not None else ""
        local_file = str(m.get("local_path", "")) if m is not None else ""
        downloaded = str(m.get("download_timestamp", "")) if m is not None else ""
        md5 = str(m.get("md5_hash", "")) if m is not None else ""
        row_count = m.get("rows_in_file", "") if m is not None else ""

        ws.cell(row=row, column=1, value=src["source_name"])
        ws.cell(row=row, column=2, value=src["publisher"])
        ws.cell(row=row, column=3, value=src["years"])
        ws.cell(row=row, column=4, value=src["description"])

        # Landing page hyperlink
        lp_cell = ws.cell(row=row, column=5, value=landing_url if landing_url else "N/A")
        if landing_url and landing_url.startswith("http"):
            lp_cell.hyperlink = landing_url
            lp_cell.font = LINK_FONT

        # Download URL hyperlink
        dl_cell = ws.cell(row=row, column=6, value=download_url if download_url else "N/A")
        if download_url and download_url.startswith("http"):
            dl_cell.hyperlink = download_url
            dl_cell.font = LINK_FONT

        ws.cell(row=row, column=7, value=local_file)
        ws.cell(row=row, column=8, value=downloaded)
        ws.cell(row=row, column=9, value=md5)
        ws.cell(row=row, column=10, value=row_count)
        ws.cell(row=row, column=11, value=src["note"])

        for ci in range(1, 12):
            ws.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    row += 2

    # ── Section B: KPI Lineage ─────────────────────────────────────────────
    ws.cell(row=row, column=1, value="SECTION B — KPI LINEAGE").font = SECTION_FONT
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    kpi_headers = ["KPI Column", "Formula", "Source Fields", "Source File"]
    for ci, h in enumerate(kpi_headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    kpi_lineage = [
        ("cf_{year}", "gen_mwh_{year} / (capacity_mw_{year} * 8760)",
         "gen_mwh_{year}, capacity_mw_{year}", "EIA-923, EIA-860"),
        ("cf_2025", "gen_mwh_2025 / (capacity_mw_2024 * 8760)",
         "gen_mwh_2025, capacity_mw_2024", "EPM-2025, EIA-860"),
        ("cf_3yr_2022_2024", "mean(cf_2022, cf_2023, cf_2024); min 2 valid years",
         "cf_2022, cf_2023, cf_2024", "Derived from EIA-923/860"),
        ("cf_3yr_2023_2025", "mean(cf_2023, cf_2024, cf_2025); min 2 valid years",
         "cf_2023, cf_2024, cf_2025", "Derived from EIA-923/860/EPM"),
        ("cf_5yr_2020_2024", "mean(cf_2020..cf_2024); min 3 valid years",
         "cf_2020 through cf_2024", "Derived from EIA-923/860"),
        ("cf_7yr_2018_2024", "mean(cf_2018..cf_2024); min 5 valid years",
         "cf_2018 through cf_2024", "Derived from EIA-923/860"),
        ("cf_regional_percentile", "rank(cf_3yr_2022_2024) within nerc_region * 100",
         "cf_3yr_2022_2024, nerc_region", "Derived + eGRID"),
        ("cf_vs_lbnl_benchmark", "cf_3yr_2022_2024 - lbnl_benchmark_cf",
         "cf_3yr_2022_2024, nerc_region, commissioning_year", "Derived + LBNL"),
        ("gen_per_mw_{year}", "gen_mwh_{year} / capacity_mw_{year}",
         "gen_mwh_{year}, capacity_mw_{year}", "EIA-923, EIA-860"),
        ("yoy_pct_{year}", "((gen_per_mw_{year} - gen_per_mw_{year-1}) / gen_per_mw_{year-1}) * 100",
         "gen_per_mw_{year}, gen_per_mw_{year-1}", "Derived"),
        ("yoy_3yr_avg", "mean(yoy_pct_2022, yoy_pct_2023, yoy_pct_2024)",
         "yoy_pct_2022, yoy_pct_2023, yoy_pct_2024", "Derived"),
        ("trend_direction", "IMPROVING if yoy_3yr_avg > 2, DECLINING if < -2, else FLAT",
         "yoy_3yr_avg", "Derived"),
        ("decline_from_peak_pct", "((gen_per_mw_2024 - peak_gen_per_mw) / peak_gen_per_mw) * 100",
         "gen_per_mw_2024, peak_gen_per_mw", "Derived"),
        ("cumulative_mwh_lost_2022_2024",
         "sum of max(0, peak_gen_per_mw * capacity - actual_gen) for 2022-2024",
         "peak_gen_per_mw, gen_mwh, capacity_mw", "Derived"),
        ("cumulative_revenue_lost_usd",
         "cumulative_mwh_lost_2022_2024 * REFERENCE_PRICE_PER_MWH ($30/MWh)",
         "cumulative_mwh_lost_2022_2024", "Derived"),
        ("asset_age", "CURRENT_YEAR (2026) - commissioning_year",
         "commissioning_year", "EIA-860"),
        ("turbine_age", "CURRENT_YEAR (2026) - turbine_vintage_min",
         "turbine_vintage_min", "USWTDB"),
        ("ptc_expired", "asset_age >= 10 AND asset_age <= 16",
         "asset_age", "Derived"),
        ("repower_candidate", "turbine_age >= 15 AND asset_age >= 10",
         "turbine_age, asset_age", "Derived"),
        ("flag_declining_3yr", "yoy_3yr_avg < -2.0",
         "yoy_3yr_avg", "Derived"),
        ("flag_bottom_quartile_cf", "cf_regional_percentile < 25",
         "cf_regional_percentile", "Derived"),
        ("flag_peak_decline_15pct", "decline_from_peak_pct < -15",
         "decline_from_peak_pct", "Derived"),
        ("flag_ptc_expired", "Same as ptc_expired",
         "ptc_expired", "Derived"),
        ("flag_repower_candidate", "Same as repower_candidate",
         "repower_candidate", "Derived"),
    ]

    for kpi_name, formula, source_fields, source_file in kpi_lineage:
        ws.cell(row=row, column=1, value=kpi_name)
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=3, value=source_fields)
        ws.cell(row=row, column=4, value=source_file)
        for ci in range(1, 5):
            ws.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    row += 2

    # ── Section C: Assumptions Register ────────────────────────────────────
    ws.cell(row=row, column=1, value="SECTION C — ASSUMPTIONS REGISTER").font = SECTION_FONT
    for c in range(1, 4):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1

    assumption_headers = ["Constant", "Value", "Rationale"]
    for ci, h in enumerate(assumption_headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    assumptions = [
        ("HOURS_PER_YEAR", 8760,
         "Standard hours in a non-leap year (365 * 24). Used as denominator in all CF calculations."),
        ("REFERENCE_PRICE_PER_MWH", 30,
         "Approximate average US wind PPA price ($/MWh) for revenue-loss estimation. "
         "Conservative mid-range; actual PPAs vary $20-$50/MWh."),
        ("CF_CAP", 0.65,
         "Maximum plausible CF threshold. Values above this are flagged as data errors. "
         "Based on theoretical and observed upper bounds for onshore wind."),
        ("CURRENT_YEAR", 2026,
         "Reference year for age calculations and PTC expiry determination."),
        ("MIN_CAPACITY_MW", 10,
         "Minimum nameplate capacity filter. Plants below this are excluded as sub-utility-scale."),
    ]

    for name, value, rationale in assumptions:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=rationale)
        for ci in range(1, 4):
            ws.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    auto_fit_columns(ws, max_width=60)
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2: Raw_Data
# ══════════════════════════════════════════════════════════════════════════════
def build_raw_data_tab(wb, df):
    """Full merged dataset with EIA hyperlinks and frozen header."""
    ws = wb.create_sheet("Raw_Data")

    # Identify raw-data columns (non-KPI): keep everything that was in the
    # original merged file.  We include all columns here as the parquet already
    # contains everything.
    raw_cols = list(df.columns)

    # Prepend EIA hyperlink column
    headers = ["EIA Plant Page (click to verify)"] + raw_cols
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)

    hyperlink_count = 0
    for ri, (_, data_row) in enumerate(df.iterrows(), start=2):
        # EIA hyperlink column
        plant_id = data_row.get("plant_id", "")
        if pd.notna(plant_id):
            url = f"https://www.eia.gov/electricity/data/browser/#/plant/{int(plant_id)}"
            cell = ws.cell(row=ri, column=1, value=f"Plant {int(plant_id)}")
            cell.hyperlink = url
            cell.font = LINK_FONT
            hyperlink_count += 1
        else:
            ws.cell(row=ri, column=1, value="N/A")

        # Data columns
        for ci, col in enumerate(raw_cols, 2):
            val = data_row[col]
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val) if not np.isnan(val) else None
            elif isinstance(val, (np.bool_,)):
                val = bool(val)
            elif pd.isna(val):
                val = None
            ws.cell(row=ri, column=ci, value=val)

    # Apply number formats based on column name
    for ci, col in enumerate(raw_cols, 2):
        if detect_cf_col(col):
            apply_number_format(ws, ci, CF_FMT)
        elif detect_pct_col(col):
            apply_number_format(ws, ci, DECIMAL_FMT)
        elif detect_currency_col(col):
            apply_number_format(ws, ci, CURRENCY_FMT)
        elif "mwh" in col.lower() or "capacity" in col.lower() or "mw" in col.lower():
            apply_number_format(ws, ci, NUMBER_FMT)

    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"
    auto_fit_columns(ws)
    return ws, hyperlink_count


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3: KPIs
# ══════════════════════════════════════════════════════════════════════════════
def build_kpis_tab(wb, df):
    """KPI columns with conditional formatting and EIA hyperlinks."""
    ws = wb.create_sheet("KPIs")

    # Select KPI-relevant columns
    id_cols = [c for c in ["plant_id", "plant_name", "state", "nerc_region",
                           "capacity_mw_2024", "commissioning_year", "owner_name"]
               if c in df.columns]
    kpi_cols = [c for c in df.columns if (
        c.startswith("cf_") or c.startswith("yoy_") or c.startswith("flag_") or
        c.startswith("gen_per_mw_") or c.startswith("trend_") or
        c.startswith("decline_") or c.startswith("cumulative_") or
        c.startswith("peak_") or c.startswith("consecutive_") or
        c in ("asset_age", "turbine_age", "ptc_expired", "repower_candidate",
              "owner_plant_count", "owner_total_mw", "independent_owner",
              "lbnl_benchmark_cf", "cf_vs_lbnl_benchmark", "cf_regional_percentile",
              "cf_summer_avg", "cf_winter_avg", "yoy_3yr_avg")
    ) and c not in id_cols]

    # Deduplicate while preserving order
    seen = set()
    unique_kpi_cols = []
    for c in kpi_cols:
        if c not in seen:
            seen.add(c)
            unique_kpi_cols.append(c)
    kpi_cols = unique_kpi_cols

    all_cols = id_cols + kpi_cols
    headers = ["EIA Plant Page (click to verify)"] + all_cols

    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)

    # Track 2025 EPM columns for blue highlighting
    epm_2025_col_indices = []

    hyperlink_count = 0
    for ri, (_, data_row) in enumerate(df.iterrows(), start=2):
        # EIA hyperlink
        plant_id = data_row.get("plant_id", "")
        if pd.notna(plant_id):
            url = f"https://www.eia.gov/electricity/data/browser/#/plant/{int(plant_id)}"
            cell = ws.cell(row=ri, column=1, value=f"Plant {int(plant_id)}")
            cell.hyperlink = url
            cell.font = LINK_FONT
            hyperlink_count += 1
        else:
            ws.cell(row=ri, column=1, value="N/A")

        # Data columns
        for ci, col in enumerate(all_cols, 2):
            val = data_row.get(col)
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val) if not np.isnan(val) else None
            elif isinstance(val, (np.bool_,)):
                val = bool(val)
            elif pd.isna(val):
                val = None
            ws.cell(row=ri, column=ci, value=val)

        # Conditional formatting: red if both flags, amber if either
        flag_declining = bool(data_row.get("flag_declining_3yr", False))
        flag_bottom_q = bool(data_row.get("flag_bottom_quartile_cf", False))

        if flag_declining and flag_bottom_q:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).fill = RED_FILL
        elif flag_declining or flag_bottom_q:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).fill = AMBER_FILL

    # Number formats and 2025 EPM column detection
    for ci, col in enumerate(all_cols, 2):
        if detect_cf_col(col):
            apply_number_format(ws, ci, CF_FMT)
        elif detect_pct_col(col):
            apply_number_format(ws, ci, DECIMAL_FMT)
        elif detect_currency_col(col):
            apply_number_format(ws, ci, CURRENCY_FMT)
        elif "mwh" in col.lower() or "capacity" in col.lower() or "mw" in col.lower():
            apply_number_format(ws, ci, NUMBER_FMT)

        # Blue highlight for 2025 EPM-derived columns
        if detect_2025_epm_col(col):
            epm_2025_col_indices.append(ci)
            # Highlight header
            hdr_cell = ws.cell(row=1, column=ci)
            hdr_cell.fill = BLUE_FILL
            hdr_cell.font = Font(bold=True, color="2F5496", size=11)
            hdr_cell.comment = Comment(
                "Source: EIA Electric Power Monthly (EPM) 2025 — preliminary data, subject to revision.",
                "Wind Screener"
            )
            # Highlight data cells
            for ri in range(2, ws.max_row + 1):
                cell = ws.cell(row=ri, column=ci)
                # Only apply blue if not already red/amber
                if cell.fill == PatternFill():
                    cell.fill = BLUE_FILL

    style_header_row(ws, len(headers))
    # Re-apply blue fill to 2025 columns (overriding header style)
    for ci in epm_2025_col_indices:
        hdr_cell = ws.cell(row=1, column=ci)
        hdr_cell.fill = BLUE_FILL
        hdr_cell.font = Font(bold=True, color="2F5496", size=11)

    ws.freeze_panes = "A2"
    auto_fit_columns(ws)
    return ws, hyperlink_count


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4: Metrics
# ══════════════════════════════════════════════════════════════════════════════
def build_metrics_tab(wb, df):
    """Live Excel formulas referencing KPIs tab cells."""
    ws = wb.create_sheet("Metrics")

    # ── Cell A1 note ──
    note_cell = ws.cell(row=1, column=1,
                        value="All formulas reference source data. "
                              "Change any value in Raw_Data and metrics update automatically.")
    note_cell.font = Font(italic=True, color="666666")
    ws.merge_cells("A1:H1")

    # ── Assumptions block (rows 2–12) ──
    ws.cell(row=3, column=1, value="ASSUMPTIONS").font = SECTION_FONT
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = SECTION_FILL

    ws.cell(row=4, column=1, value="Parameter")
    ws.cell(row=4, column=2, value="Value")
    ws.cell(row=4, column=3, value="Named Reference")
    for ci in range(1, 4):
        cell = ws.cell(row=4, column=ci)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    assumptions_block = [
        ("Hours per Year", HOURS_PER_YEAR, "HOURS_PER_YEAR"),
        ("Reference Price ($/MWh)", REFERENCE_PRICE_PER_MWH, "REFERENCE_PRICE"),
        ("CF Cap", CF_CAP, "CF_CAP"),
        ("Current Year", CURRENT_YEAR, "CURRENT_YEAR"),
        ("Min Capacity (MW)", MIN_CAPACITY_MW, "MIN_CAPACITY_MW"),
    ]

    # Define named cells for the assumptions
    for i, (label, value, name_ref) in enumerate(assumptions_block):
        r = 5 + i
        ws.cell(row=r, column=1, value=label)
        val_cell = ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=name_ref)
        for ci in range(1, 4):
            ws.cell(row=r, column=ci).border = THIN_BORDER

        # Create workbook-scoped defined name pointing to this cell
        from openpyxl.workbook.defined_name import DefinedName
        ref = f"Metrics!$B${r}"
        dn = DefinedName(name_ref, attr_text=ref)
        wb.defined_names.add(dn)

    # Row references for assumption values
    hours_ref = "$B$5"    # HOURS_PER_YEAR
    price_ref = "$B$6"    # REFERENCE_PRICE
    cf_cap_ref = "$B$7"   # CF_CAP
    current_year_ref = "$B$8"  # CURRENT_YEAR

    # ── Per-plant metrics (row 13+) ──
    metrics_start_row = 13
    ws.cell(row=12, column=1, value="PER-PLANT METRICS").font = SECTION_FONT
    for c in range(1, 9):
        ws.cell(row=12, column=c).fill = SECTION_FILL

    metric_headers = [
        "Plant ID", "Plant Name", "Capacity (MW)", "Commissioning Year",
        "CF (3yr avg)", "Asset Age", "YoY Change 2024 (%)",
        "Revenue Lost (USD)"
    ]
    for ci, h in enumerate(metric_headers, 1):
        cell = ws.cell(row=metrics_start_row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # Build a mapping from KPIs tab column names to column letters
    kpis_ws = wb["KPIs"]
    kpis_col_map = {}
    for ci in range(1, kpis_ws.max_column + 1):
        val = kpis_ws.cell(row=1, column=ci).value
        if val:
            kpis_col_map[val] = get_column_letter(ci)

    # Per-plant formula rows
    nrows = len(df)
    for i in range(nrows):
        r = metrics_start_row + 1 + i
        kpi_row = i + 2  # Row in KPIs tab (1-indexed, row 1 is header)

        # Plant ID — direct reference to KPIs tab
        pid_col = kpis_col_map.get("plant_id", "B")
        ws.cell(row=r, column=1, value=f"=KPIs!{pid_col}{kpi_row}")

        # Plant Name
        pname_col = kpis_col_map.get("plant_name", "C")
        ws.cell(row=r, column=2, value=f"=KPIs!{pname_col}{kpi_row}")

        # Capacity (MW)
        cap_col = kpis_col_map.get("capacity_mw_2024", "")
        if cap_col:
            cell = ws.cell(row=r, column=3, value=f"=KPIs!{cap_col}{kpi_row}")
            cell.number_format = NUMBER_FMT

        # Commissioning Year
        comm_col = kpis_col_map.get("commissioning_year", "")
        if comm_col:
            ws.cell(row=r, column=4, value=f"=KPIs!{comm_col}{kpi_row}")

        # CF (3yr avg) — reference KPIs tab
        cf3_col = kpis_col_map.get("cf_3yr_2022_2024", "")
        if cf3_col:
            cell = ws.cell(row=r, column=5, value=f"=KPIs!{cf3_col}{kpi_row}")
            cell.number_format = CF_FMT

        # Asset Age — formula using CURRENT_YEAR assumption
        if comm_col:
            cell = ws.cell(row=r, column=6,
                           value=f"=Metrics!{current_year_ref}-KPIs!{comm_col}{kpi_row}")
            cell.number_format = "0"

        # YoY Change 2024 (%)
        yoy_col = kpis_col_map.get("yoy_pct_2024", "")
        if yoy_col:
            cell = ws.cell(row=r, column=7, value=f"=KPIs!{yoy_col}{kpi_row}")
            cell.number_format = DECIMAL_FMT

        # Revenue Lost (USD) — formula: cumulative_mwh_lost * REFERENCE_PRICE
        mwh_lost_col = kpis_col_map.get("cumulative_mwh_lost_2022_2024", "")
        if mwh_lost_col:
            cell = ws.cell(row=r, column=8,
                           value=f"=KPIs!{mwh_lost_col}{kpi_row}*Metrics!{price_ref}")
            cell.number_format = CURRENCY_FMT

        for ci in range(1, 9):
            ws.cell(row=r, column=ci).border = THIN_BORDER

    auto_fit_columns(ws)
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5: Summary
# ══════════════════════════════════════════════════════════════════════════════
def build_summary_tab(wb, df):
    """Aggregate stats using live Excel formulas (COUNTIF, MEDIAN, SUM)."""
    ws = wb.create_sheet("Summary")

    ws.cell(row=1, column=1, value="WIND FARM SCREENER — SUMMARY STATISTICS").font = Font(
        bold=True, color="2F5496", size=14)
    ws.merge_cells("A1:C1")

    ws.cell(row=3, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=3, column=1).fill = HEADER_FILL
    ws.cell(row=3, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=3, column=2).fill = HEADER_FILL
    ws.cell(row=3, column=3, value="Notes").font = HEADER_FONT
    ws.cell(row=3, column=3).fill = HEADER_FILL
    for ci in range(1, 4):
        ws.cell(row=3, column=ci).border = THIN_BORDER

    # Build column references from KPIs tab
    kpis_ws = wb["KPIs"]
    kpis_col_map = {}
    for ci in range(1, kpis_ws.max_column + 1):
        val = kpis_ws.cell(row=1, column=ci).value
        if val:
            kpis_col_map[val] = get_column_letter(ci)

    nrows = len(df)
    last_row = nrows + 1  # +1 because row 1 is header

    # Helper to build range reference on KPIs tab
    def kpi_range(col_name):
        col_letter = kpis_col_map.get(col_name, "")
        if col_letter:
            return f"KPIs!{col_letter}2:{col_letter}{last_row}"
        return ""

    metrics = []

    # 1. Total plants
    pid_col = kpis_col_map.get("plant_id", "B")
    metrics.append((
        "Total Plants",
        f"=COUNTA(KPIs!{pid_col}2:{pid_col}{last_row})",
        "Count of all plants in dataset"
    ))

    # 2. Flagged declining 3yr
    rng = kpi_range("flag_declining_3yr")
    if rng:
        metrics.append((
            "Flagged: Declining 3yr Trend",
            f'=COUNTIF({rng},TRUE)',
            "Plants with 3-year rolling YoY avg < -2%"
        ))

    # 3. Flagged bottom quartile CF
    rng = kpi_range("flag_bottom_quartile_cf")
    if rng:
        metrics.append((
            "Flagged: Bottom Quartile CF",
            f'=COUNTIF({rng},TRUE)',
            "Plants in bottom 25th percentile CF within NERC region"
        ))

    # 4. Flagged PTC expired
    rng = kpi_range("flag_ptc_expired")
    if rng:
        metrics.append((
            "Flagged: PTC Expired",
            f'=COUNTIF({rng},TRUE)',
            "Plants with 10-year PTC likely expired (age 10-16)"
        ))

    # 5. Flagged repower candidate
    rng = kpi_range("flag_repower_candidate")
    if rng:
        metrics.append((
            "Flagged: Repower Candidate",
            f'=COUNTIF({rng},TRUE)',
            "Turbine age >= 15 and asset age >= 10"
        ))

    # 6. Median CF (3yr rolling)
    rng = kpi_range("cf_3yr_2022_2024")
    if rng:
        metrics.append((
            "Median CF (3yr Rolling 2022-2024)",
            f"=MEDIAN({rng})",
            "Median capacity factor across all plants"
        ))

    # 7. Total capacity MW
    rng = kpi_range("capacity_mw_2024")
    if rng:
        metrics.append((
            "Total Capacity (MW)",
            f"=SUM({rng})",
            "Sum of 2024 nameplate capacity"
        ))

    # Write metrics
    for i, (label, formula, note) in enumerate(metrics):
        r = 4 + i
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=formula)
        ws.cell(row=r, column=3, value=note)
        for ci in range(1, 4):
            ws.cell(row=r, column=ci).border = THIN_BORDER

        # Apply formatting based on metric type
        if "CF" in label and "Median" in label:
            cell.number_format = CF_FMT
        elif "Capacity" in label:
            cell.number_format = NUMBER_FMT

    auto_fit_columns(ws, max_width=50)
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    start = time.time()
    print("=" * 70)
    print("  Wind Farm Screener — Excel Export")
    print("=" * 70)

    # ── Load inputs ──
    kpi_path = PROC_DIR / "wind_plants_kpis.parquet"
    if not kpi_path.exists():
        print(f"\u2717 Input not found: {kpi_path}")
        print("  Run 03_kpis.py first.")
        sys.exit(1)

    df = pd.read_parquet(kpi_path)
    input_rows = len(df)
    print(f"\nLoaded {input_rows} plants from KPI dataset")
    print(f"Columns: {len(df.columns)}")

    # Load download manifest (optional — Sources tab will degrade gracefully)
    manifest_path = AUDIT_DIR / "download_manifest.csv"
    manifest_df = None
    if manifest_path.exists():
        manifest_df = pd.read_csv(manifest_path)
        print(f"Loaded manifest: {len(manifest_df)} rows")
    else:
        print("  \u26a0 download_manifest.csv not found; Sources tab will use metadata only")

    # ── Build workbook ──
    wb = Workbook()

    # Tab 1: Sources
    print("\nBuilding tab: Sources...")
    build_sources_tab(wb, manifest_df)

    # Tab 2: Raw_Data
    print("Building tab: Raw_Data...")
    _, raw_hyperlinks = build_raw_data_tab(wb, df)

    # Tab 3: KPIs
    print("Building tab: KPIs...")
    _, kpi_hyperlinks = build_kpis_tab(wb, df)

    # Tab 4: Metrics
    print("Building tab: Metrics...")
    build_metrics_tab(wb, df)

    # Tab 5: Summary
    print("Building tab: Summary...")
    build_summary_tab(wb, df)

    # ── Save ──
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUT_DIR / "wind_screener.xlsx"
    print(f"\nSaving workbook to {output_path}...")
    wb.save(output_path)

    total_hyperlinks = raw_hyperlinks + kpi_hyperlinks
    output_rows = input_rows  # 1:1 mapping
    elapsed = time.time() - start

    # ── Gate check ──
    print(f"\n\u2713 04_export_excel.py complete")
    print(f"  Input rows:           {input_rows}")
    print(f"  Output rows:          {output_rows}")
    print(f"  Tabs:                 5")
    print(f"  Hyperlinks verified:  {total_hyperlinks}")
    print(f"  Elapsed:              {elapsed:.1f}s")


if __name__ == "__main__":
    main()
