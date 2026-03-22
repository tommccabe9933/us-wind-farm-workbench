"""
05_verify.py — Validation and audit checks for the Wind Farm Investment Screener.

Run after all other scripts (01–04).

Input:
  outputs/audit/download_manifest.csv
  data/processed/wind_plants_kpis.parquet
  outputs/wind_plants.csv
  outputs/wind_screener.xlsx
  outputs/audit/source_trace.csv

Output:
  outputs/audit/validation_report.txt
"""

import hashlib
import time
import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
import numpy as np

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent
PROC_DIR = BASE_DIR / "data" / "processed"
OUT_DIR = BASE_DIR / "outputs"
AUDIT_DIR = OUT_DIR / "audit"

MANIFEST_PATH = AUDIT_DIR / "download_manifest.csv"
KPI_PATH = PROC_DIR / "wind_plants_kpis.parquet"
CSV_PATH = OUT_DIR / "wind_plants.csv"
XLSX_PATH = OUT_DIR / "wind_screener.xlsx"
TRACE_PATH = AUDIT_DIR / "source_trace.csv"

# ── Reference constants ───────────────────────────────────────────────────────
EIA_2024_WIND_TWH = 451  # EIA-published U.S. wind total for 2024
EIA_2025_WIND_TWH = 464  # EIA-published U.S. wind total for 2025


def md5_file(path):
    """Compute MD5 hash for a file."""
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def check_1_file_integrity(manifest):
    """Re-compute MD5 for each raw file listed in download_manifest.csv."""
    status = "PASS"
    detail = ""
    mismatches = []
    missing = []

    for _, row in manifest.iterrows():
        local_path = row.get("local_path", "")
        stored_hash = row.get("md5_hash", "")

        if not local_path or pd.isna(local_path):
            continue
        if not stored_hash or pd.isna(stored_hash):
            continue

        fpath = Path(local_path)
        if not fpath.is_absolute():
            fpath = BASE_DIR / fpath

        if not fpath.exists():
            missing.append(str(fpath))
            continue

        actual_hash = md5_file(fpath)
        if actual_hash != stored_hash:
            mismatches.append(str(fpath))

    if mismatches:
        status = "FAIL"
        detail = f"{len(mismatches)} hash mismatch(es)"
    elif missing:
        status = "FAIL"
        detail = f"{len(missing)} file(s) missing"
    else:
        detail = "all hashes verified"

    return status, detail


def check_2_plant_count(df):
    """Total >=10 MW wind plants should be 900-2100. Total capacity 120k-175k MW."""
    n_plants = len(df)
    total_cap = df["capacity_mw_2024"].sum() if "capacity_mw_2024" in df.columns else 0

    if 900 <= n_plants <= 2100 and 120_000 <= total_cap <= 175_000:
        status = "PASS"
    elif n_plants < 900 or n_plants > 2100:
        status = "WARN" if 700 <= n_plants <= 2500 else "FAIL"
    elif total_cap < 120_000 or total_cap > 175_000:
        status = "WARN" if 100_000 <= total_cap <= 200_000 else "FAIL"
    else:
        status = "WARN"

    detail = f"{n_plants} plants, {total_cap:,.0f} MW"
    return status, detail


def check_3_gen_2024(df):
    """Sum gen_mwh_2024 and compare to EIA-published ~451 TWh."""
    if "gen_mwh_2024" not in df.columns:
        return "FAIL", "gen_mwh_2024 column missing"

    total_mwh = df["gen_mwh_2024"].sum()
    eia_total_mwh = EIA_2024_WIND_TWH * 1e6
    pct = (total_mwh / eia_total_mwh) * 100

    print(f"  Dataset 2024 total: {total_mwh:,.0f} MWh ({pct:.1f}% of ~{EIA_2024_WIND_TWH} TWh)")

    if pct < 70:
        status = "FAIL"
    elif 85 <= pct <= 95:
        status = "PASS"
    else:
        status = "WARN"

    detail = f"{pct:.1f}% coverage"
    return status, detail


def check_4_gen_2025(df):
    """Sum gen_mwh_2025 and compare to EIA-published ~464 TWh."""
    if "gen_mwh_2025" not in df.columns:
        return "FAIL", "gen_mwh_2025 column missing"

    total_mwh = df["gen_mwh_2025"].sum()
    eia_total_mwh = EIA_2025_WIND_TWH * 1e6
    pct = (total_mwh / eia_total_mwh) * 100

    print(f"  Dataset 2025 total: {total_mwh:,.0f} MWh ({pct:.1f}% of ~{EIA_2025_WIND_TWH} TWh)")

    if pct < 70:
        status = "FAIL"
    elif 85 <= pct <= 95:
        status = "PASS"
    else:
        status = "WARN"

    detail = f"{pct:.1f}% coverage"
    return status, detail


def check_5_cf_range(df):
    """No cf_3yr_2022_2024 > 0.65. Print median."""
    col = "cf_3yr_2022_2024"
    if col not in df.columns:
        return "FAIL", "cf_3yr_2022_2024 column missing"

    vals = df[col].dropna()
    above_cap = (vals > 0.65).sum()
    median_cf = vals.median()

    if above_cap > 0:
        status = "FAIL"
        detail = f"median={median_cf:.4f}, {above_cap} plants > 0.65"
    else:
        status = "PASS"
        detail = f"median={median_cf:.4f}"

    return status, detail


def check_6_rolling_avg(df):
    """Verify cf_3yr_2022_2024 equals mean(cf_2022, cf_2023, cf_2024) within 0.001."""
    required = ["cf_3yr_2022_2024", "cf_2022", "cf_2023", "cf_2024"]
    for c in required:
        if c not in df.columns:
            return "FAIL", f"{c} column missing"

    subset = df[df["cf_3yr_2022_2024"].notna()].copy()
    if len(subset) == 0:
        return "FAIL", "no non-null cf_3yr_2022_2024 values"

    # Sample up to 100 random plants
    n_sample = min(100, len(subset))
    sample = subset.sample(n=n_sample, random_state=42)

    computed = sample[["cf_2022", "cf_2023", "cf_2024"]].mean(axis=1)
    stored = sample["cf_3yr_2022_2024"]
    diff = (computed - stored).abs()
    failures = (diff > 0.001).sum()

    if failures > 0:
        return "FAIL", f"{failures}/{n_sample} sampled plants exceed 0.001 tolerance"

    return "PASS", f"{n_sample} sampled plants verified"


def check_7_hyperlinks(xlsx_path):
    """Check first 10 rows of KPIs sheet for valid hyperlink objects."""
    try:
        import openpyxl
    except ImportError:
        return "FAIL", "openpyxl not installed"

    import re

    if not xlsx_path.exists():
        return "FAIL", f"{xlsx_path.name} not found"

    wb = openpyxl.load_workbook(xlsx_path)

    # Find KPIs sheet (case-insensitive search)
    kpi_sheet = None
    for name in wb.sheetnames:
        if "kpi" in name.lower():
            kpi_sheet = wb[name]
            break

    if kpi_sheet is None:
        return "FAIL", "no KPIs sheet found"

    pattern = re.compile(r"https://www\.eia\.gov/electricity/data/browser/#/plant/\d+")
    found = 0
    checked = 0

    for row_idx in range(2, min(12, kpi_sheet.max_row + 1)):  # rows 2–11 (skip header)
        for cell in kpi_sheet[row_idx]:
            if cell.hyperlink and cell.hyperlink.target:
                checked += 1
                if pattern.match(cell.hyperlink.target):
                    found += 1

    if found == 0 and checked == 0:
        return "FAIL", "no hyperlink objects found in first 10 rows"
    elif found == 0:
        return "FAIL", f"{checked} hyperlinks checked, none match expected pattern"

    return "PASS", f"{found} valid hyperlinks in first 10 rows"


def check_8_source_trace(csv_path, trace_path):
    """Every plant in wind_plants.csv has an entry in source_trace.csv with non-null eia_plant_url."""
    if not csv_path.exists():
        return "FAIL", f"{csv_path.name} not found"
    if not trace_path.exists():
        return "FAIL", f"{trace_path.name} not found"

    plants = pd.read_csv(csv_path)
    trace = pd.read_csv(trace_path)

    if "eia_plant_id" not in plants.columns:
        # Try alternate column name
        plant_id_col = [c for c in plants.columns if "plant_id" in c.lower()]
        if not plant_id_col:
            return "FAIL", "no plant_id column found in wind_plants.csv"
        pid_col = plant_id_col[0]
    else:
        pid_col = "eia_plant_id"

    if "eia_plant_id" not in trace.columns:
        trace_id_col = [c for c in trace.columns if "plant_id" in c.lower()]
        if not trace_id_col:
            return "FAIL", "no plant_id column found in source_trace.csv"
        trace_pid_col = trace_id_col[0]
    else:
        trace_pid_col = "eia_plant_id"

    plant_ids = set(plants[pid_col].dropna().astype(int))
    trace_ids = set(trace[trace_pid_col].dropna().astype(int))

    missing_from_trace = plant_ids - trace_ids
    if missing_from_trace:
        return "FAIL", f"{len(missing_from_trace)} plants missing from source_trace.csv"

    # Check non-null eia_plant_url
    if "eia_plant_url" not in trace.columns:
        return "FAIL", "eia_plant_url column missing from source_trace.csv"

    trace_for_plants = trace[trace[trace_pid_col].isin(plant_ids)]
    null_urls = trace_for_plants["eia_plant_url"].isna().sum()
    if null_urls > 0:
        return "FAIL", f"{null_urls} plants with null eia_plant_url"

    return "PASS", f"all {len(plant_ids)} plants have source trace entries"


def check_9_flag_logic(df):
    """Verify flag logic consistency."""
    errors = []

    # flag_consec_decline_3yr = True requires consecutive_decline_years >= 3
    if "flag_consec_decline_3yr" in df.columns and "consecutive_decline_years" in df.columns:
        flagged = df[df["flag_consec_decline_3yr"] == True]
        bad = flagged[flagged["consecutive_decline_years"] < 3]
        if len(bad) > 0:
            errors.append(f"flag_consec_decline_3yr: {len(bad)} plants with consecutive_decline_years < 3")

    # flag_ptc_expired = True requires 10 <= asset_age <= 16
    if "flag_ptc_expired" in df.columns and "asset_age" in df.columns:
        flagged = df[df["flag_ptc_expired"] == True]
        bad = flagged[(flagged["asset_age"] < 10) | (flagged["asset_age"] > 16)]
        if len(bad) > 0:
            errors.append(f"flag_ptc_expired: {len(bad)} plants outside 10-16 asset_age range")

    # flag_repower_candidate = True requires turbine_age >= 15 AND asset_age >= 10
    if "flag_repower_candidate" in df.columns:
        flagged = df[df["flag_repower_candidate"] == True]
        if "turbine_age" in df.columns and "asset_age" in df.columns:
            bad = flagged[(flagged["turbine_age"] < 15) | (flagged["asset_age"] < 10)]
            if len(bad) > 0:
                errors.append(f"flag_repower_candidate: {len(bad)} plants fail turbine_age>=15 AND asset_age>=10")

    # flag_declining_3yr = True must have non-null yoy_3yr_avg
    if "flag_declining_3yr" in df.columns and "yoy_3yr_avg" in df.columns:
        flagged = df[df["flag_declining_3yr"] == True]
        null_yoy = flagged["yoy_3yr_avg"].isna().sum()
        if null_yoy > 0:
            errors.append(f"flag_declining_3yr: {null_yoy} plants with null yoy_3yr_avg")

    if errors:
        return "FAIL", "; ".join(errors)

    return "PASS", "all flag logic consistent"


def main():
    t0 = time.time()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    AUDIT_DIR.mkdir(parents=True, exist_ok=True)

    # ── Load data ──────────────────────────────────────────────────────────────
    manifest = pd.read_csv(MANIFEST_PATH) if MANIFEST_PATH.exists() else None
    df = pd.read_parquet(KPI_PATH) if KPI_PATH.exists() else None

    if manifest is None:
        print("ERROR: download_manifest.csv not found. Run 01_download.py first.")
        sys.exit(1)
    if df is None:
        print("ERROR: wind_plants_kpis.parquet not found. Run 03_kpis.py first.")
        sys.exit(1)

    # ── Run checks ─────────────────────────────────────────────────────────────
    results = {}

    print("\nRunning CHECK 1 — File integrity...")
    results[1] = check_1_file_integrity(manifest)

    print("Running CHECK 2 — Plant count...")
    results[2] = check_2_plant_count(df)

    print("Running CHECK 3 — 2024 generation total...")
    results[3] = check_3_gen_2024(df)

    print("Running CHECK 4 — 2025 EPM total...")
    results[4] = check_4_gen_2025(df)

    print("Running CHECK 5 — CF range...")
    results[5] = check_5_cf_range(df)

    print("Running CHECK 6 — Rolling avg consistency...")
    results[6] = check_6_rolling_avg(df)

    print("Running CHECK 7 — Hyperlinks...")
    results[7] = check_7_hyperlinks(XLSX_PATH)

    print("Running CHECK 8 — Source trace...")
    results[8] = check_8_source_trace(CSV_PATH, TRACE_PATH)

    print("Running CHECK 9 — Flag logic...")
    results[9] = check_9_flag_logic(df)

    # ── Build report ───────────────────────────────────────────────────────────
    check_labels = {
        1: "File integrity",
        2: "Plant count",
        3: "2024 generation total",
        4: "2025 EPM total",
        5: "CF range",
        6: "Rolling avg consistency",
        7: "Hyperlinks",
        8: "Source trace",
        9: "Flag logic",
    }

    has_fail = any(r[0] == "FAIL" for r in results.values())
    overall = "FAIL" if has_fail else "PASS"
    n_pass = sum(1 for r in results.values() if r[0] in ("PASS", "WARN"))
    n_warn = sum(1 for r in results.values() if r[0] == "WARN")

    lines = []
    lines.append("=== Wind Farm Screener Validation Report ===")
    lines.append(f"Generated:      {timestamp}")
    lines.append("Data coverage:  EIA-860/923 annual final 2018-2024 | EPM monthly Jan-Dec 2025")
    lines.append("")

    for i in range(1, 10):
        status, detail = results[i]
        label = check_labels[i]
        # Pad label to align columns
        check_str = f"CHECK {i} — {label}:"
        lines.append(f"{check_str:<42}{status:<10}{detail}")

    lines.append("")
    lines.append(f"OVERALL: {overall}")

    report = "\n".join(lines)

    # Print to console
    print("\n" + report)

    # Write to file
    report_path = AUDIT_DIR / "validation_report.txt"
    report_path.write_text(report + "\n", encoding="utf-8")

    elapsed = time.time() - t0
    print(f"\n\u2713 05_verify.py complete")
    print(f"  Checks passed:  {n_pass}/9")
    print(f"  Warnings:       {n_warn}")
    print(f"  Elapsed:        {elapsed:.1f}s")


if __name__ == "__main__":
    main()
